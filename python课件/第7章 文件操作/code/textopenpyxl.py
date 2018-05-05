import openpyxl
from openpyxl import Workbook

fn = r'f:\test.xlsx'
wb = Workbook()
ws = wb.create_sheet(title='你好，世界')
ws['A1'] = '这是第一个单元格'
ws['B1'] = 3.1415926

wb.save(fn)

wb = openpyxl.load_workbook(fn)
ws = wb.worksheets[1]
print(ws['A1'].value)

ws.append([1,2,3,4,5])
ws.merge_cells('F2:F3')
ws['F2'] = "=sum(A2:E2)"

for r in range(10,15):
    for c in range(3,8):
        _ = ws.cell(row=r, column=c, value=r*c)
wb.save(fn)
